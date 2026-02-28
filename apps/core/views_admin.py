"""Vistas CRUD del panel de administración (sin Django Admin)."""
import io
from django.db import models
from django.db.models import Q, Count, Prefetch
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator

from apps.products.models import Product, Category, Brand, ProductAttribute, ProductReview
from apps.orders.models import Order, OrderItem, OrderNote
from apps.coupons.models import Coupon
from apps.accounts.models import User
from .forms import (
    CategoryForm, BrandForm, ProductForm, OrderStatusForm, CouponForm,
    ProductAttributeForm, ProductAttributeValueFormSet,
    get_product_variant_formset, CustomerForm, CustomerCreateForm, SiteSettingsForm,
    HomeSectionForm, HomeHeroSlideForm, HomeAboutBlockForm, HomeMeatCategoryBlockForm, HomeBrandBlockForm, HomeBrandForm, HomeTestimonialForm,
    HomePopupAnnouncementForm,
    ShippingPriceForm, ShippingFreeRuleForm,
)
from .models import (
    SiteSettings,
    HomeSection,
    HomeHeroSlide,
    HomeAboutBlock,
    HomeMeatCategoryBlock,
    HomeBrandBlock,
    HomeBrand,
    HomeTestimonial,
    HomePopupAnnouncement,
    Country,
    State,
    City,
    ShippingPrice,
    NewsletterSubscriber,
    ContactSubmission,
)


class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Mixin que requiere rol staff o admin (acceso al panel de administración)."""

    def test_func(self):
        return getattr(self.request.user, 'can_access_dashboard', False)


def _safe_next_url(request, candidate, fallback):
    if candidate and url_has_allowed_host_and_scheme(
        url=candidate,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return candidate
    return fallback


def _dashboard_required(view):
    """Decorador: requiere can_access_dashboard (staff o admin)."""
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not getattr(request.user, 'can_access_dashboard', False):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Sin acceso')
        return view(request, *args, **kwargs)
    return wrapped


# --- Categorías ---

class CategoryListView(StaffRequiredMixin, ListView):
    model = Category
    template_name = 'dashboard/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20

    def get_queryset(self):
        qs = Category.objects.annotate(
            product_count=Count('products')
        )
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                models.Q(name__icontains=search) |
                models.Q(slug__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        sort = self.request.GET.get('sort', 'order')
        order_map = {
            'order': ['order', 'name'],
            'name': ['name'],
            '-name': ['-name'],
            'products': ['product_count', 'name'],
            '-products': ['-product_count', 'name'],
        }
        return qs.order_by(*order_map.get(sort, ['order', 'name']))


class CategoryProductListView(StaffRequiredMixin, ListView):
    """Lista los productos de una categoría (al hacer clic en el contador)."""
    model = Product
    template_name = 'dashboard/category_products.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        self.category = get_object_or_404(Category, pk=self.kwargs['pk'])
        return self.category.products.all().select_related('brand').order_by('name')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['category'] = self.category
        return ctx


class CategoryCreateView(StaffRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'dashboard/category_form.html'
    success_url = reverse_lazy('core:admin_panel:category_list')

    def form_valid(self, form):
        messages.success(self.request, 'Categoría creada correctamente.')
        return super().form_valid(form)


class CategoryUpdateView(StaffRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'dashboard/category_form.html'
    context_object_name = 'category'
    success_url = reverse_lazy('core:admin_panel:category_list')

    def form_valid(self, form):
        messages.success(self.request, 'Categoría actualizada correctamente.')
        return super().form_valid(form)


class CategoryDeleteView(StaffRequiredMixin, DeleteView):
    model = Category
    template_name = 'dashboard/category_confirm_delete.html'
    context_object_name = 'category'
    success_url = reverse_lazy('core:admin_panel:category_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Categoría eliminada.')
        return super().delete(request, *args, **kwargs)


@_dashboard_required
def category_toggle_active_view(request, pk):
    """Activa o inactiva una categoría."""
    if request.method != 'POST':
        return redirect('core:admin_panel:category_list')
    category = get_object_or_404(Category, pk=pk)
    category.is_active = not category.is_active
    category.save()
    action = 'activada' if category.is_active else 'inactivada'
    messages.success(request, f'Categoría «{category.name}» {action} correctamente.')
    return redirect('core:admin_panel:category_list')


# --- Marcas (catálogo) ---

class BrandListView(StaffRequiredMixin, ListView):
    model = Brand
    template_name = 'dashboard/brand_list.html'
    context_object_name = 'brands'
    paginate_by = 20

    def get_queryset(self):
        qs = Brand.objects.annotate(
            product_count=models.Count('products')
        )
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                models.Q(name__icontains=search) |
                models.Q(slug__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        sort = self.request.GET.get('sort', 'order')
        order_map = {
            'order': ['order', 'name'],
            'name': ['name'],
            '-name': ['-name'],
            'products': ['product_count', 'name'],
            '-products': ['-product_count', 'name'],
        }
        return qs.order_by(*order_map.get(sort, ['order', 'name']))


class BrandCreateView(StaffRequiredMixin, CreateView):
    model = Brand
    form_class = BrandForm
    template_name = 'dashboard/brand_form.html'
    success_url = reverse_lazy('core:admin_panel:brand_list')

    def form_valid(self, form):
        messages.success(self.request, 'Marca creada correctamente.')
        return super().form_valid(form)


class BrandUpdateView(StaffRequiredMixin, UpdateView):
    model = Brand
    form_class = BrandForm
    template_name = 'dashboard/brand_form.html'
    context_object_name = 'brand'
    success_url = reverse_lazy('core:admin_panel:brand_list')

    def form_valid(self, form):
        messages.success(self.request, 'Marca actualizada correctamente.')
        return super().form_valid(form)


class BrandDeleteView(StaffRequiredMixin, DeleteView):
    model = Brand
    template_name = 'dashboard/brand_confirm_delete.html'
    context_object_name = 'brand'
    success_url = reverse_lazy('core:admin_panel:brand_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Marca eliminada.')
        return super().delete(request, *args, **kwargs)


# --- Atributos ---

class AttributeListView(StaffRequiredMixin, ListView):
    model = ProductAttribute
    template_name = 'dashboard/attribute_list.html'
    context_object_name = 'attributes'
    paginate_by = 20

    def get_queryset(self):
        qs = ProductAttribute.objects.prefetch_related('values')
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                models.Q(name__icontains=search) |
                models.Q(slug__icontains=search)
            )
        sort = self.request.GET.get('sort', 'order')
        order_map = {
            'order': ['order', 'name'],
            'name': ['name'],
            '-name': ['-name'],
        }
        return qs.order_by(*order_map.get(sort, ['order', 'name']))


class AttributeCreateView(StaffRequiredMixin, CreateView):
    model = ProductAttribute
    form_class = ProductAttributeForm
    template_name = 'dashboard/attribute_form.html'
    success_url = reverse_lazy('core:admin_panel:attribute_list')

    def form_valid(self, form):
        messages.success(self.request, 'Atributo creado. Agrega sus valores editándolo.')
        return super().form_valid(form)


class AttributeUpdateView(StaffRequiredMixin, UpdateView):
    model = ProductAttribute
    form_class = ProductAttributeForm
    template_name = 'dashboard/attribute_form.html'
    context_object_name = 'attribute'
    success_url = reverse_lazy('core:admin_panel:attribute_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.method == 'POST':
            ctx['value_formset'] = ProductAttributeValueFormSet(
                self.request.POST, instance=self.object
            )
        else:
            ctx['value_formset'] = ProductAttributeValueFormSet(instance=self.object)
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data(form=form)
        value_formset = ctx['value_formset']
        if value_formset.is_valid():
            form.save()
            value_formset.save()
            messages.success(self.request, 'Atributo y valores actualizados.')
            return redirect(self.success_url)
        return self.render_to_response(ctx)


class AttributeDeleteView(StaffRequiredMixin, DeleteView):
    model = ProductAttribute
    template_name = 'dashboard/attribute_confirm_delete.html'
    context_object_name = 'attribute'
    success_url = reverse_lazy('core:admin_panel:attribute_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Atributo eliminado.')
        return super().delete(request, *args, **kwargs)


# --- Productos ---

class ProductListView(StaffRequiredMixin, ListView):
    model = Product
    template_name = 'dashboard/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        qs = Product.objects.select_related('brand').prefetch_related('categories', 'images')
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(codigo__icontains=search) |
                Q(short_description__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        category = self.request.GET.get('category')
        if category:
            qs = qs.filter(categories__id=category).distinct()
        brand = self.request.GET.get('brand')
        if brand:
            qs = qs.filter(brand_id=brand)
        product_type = self.request.GET.get('type')
        if product_type:
            qs = qs.filter(product_type=product_type)
        featured = self.request.GET.get('featured')
        if featured == '1':
            qs = qs.filter(is_featured=True)
        elif featured == '0':
            qs = qs.filter(is_featured=False)
        sort = self.request.GET.get('sort', '-created_at')
        allowed_sort = {'name': 'name', '-name': '-name', 'price': 'regular_price', '-price': '-regular_price',
                       'created': '-created_at', '-created': 'created_at', 'sku': 'sku'}
        order = allowed_sort.get(sort, '-created_at')
        return qs.order_by(order)

    def get_context_data(self, **kwargs):
        from apps.products.models import Category, Brand
        ctx = super().get_context_data(**kwargs)
        ctx['categories'] = Category.objects.filter(is_active=True).order_by('name')
        ctx['brands'] = Brand.objects.filter(is_active=True).order_by('name')
        return ctx


@_dashboard_required
def sync_tersa_products_view(request):
    """Importa productos desde API Tersa (solo BARBERSHOP y BARBER UP)."""
    if request.method != 'POST':
        return redirect('core:admin_panel:product_list')
    try:
        from apps.integrations.services import sync_tersa_products
        result = sync_tersa_products(brands=['BARBERSHOP', 'BARBER UP'], download_images=True)
        messages.success(
            request,
            f'Tersa: {result["total"]} productos (BARBERSHOP, BARBER UP + IDs extra). '
            f'{result["created"]} creados, {result["updated"]} actualizados.'
        )
    except Exception as e:
        messages.error(request, f'Error al sincronizar Tersa: {e}')
    return redirect('core:admin_panel:product_list')


class ProductCreateView(StaffRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'dashboard/product_form.html'
    success_url = reverse_lazy('core:admin_panel:product_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Producto creado. Edita el producto para agregar variantes si es variable.')
        return redirect('core:admin_panel:product_edit', pk=self.object.pk)


class ProductUpdateView(StaffRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'dashboard/product_form.html'
    context_object_name = 'product'
    success_url = reverse_lazy('core:admin_panel:product_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        product = self.object
        if product and product.product_type == 'variable':
            ctx['variant_formset'] = get_product_variant_formset(
                product,
                data=self.request.POST if self.request.method == 'POST' else None,
                files=self.request.FILES if self.request.method == 'POST' else None
            )
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        product = self.object
        if product.product_type == 'variable':
            formset = get_product_variant_formset(product, data=self.request.POST, files=self.request.FILES)
            if formset.is_valid():
                formset.save()
                messages.success(self.request, 'Producto y variantes actualizados.')
            else:
                messages.warning(self.request, 'Producto guardado pero hay errores en las variantes.')
        else:
            product.variants.all().delete()
            messages.success(self.request, 'Producto actualizado correctamente.')
        return response


@_dashboard_required
def product_toggle_active_view(request, pk):
    """Activa o inactiva un producto."""
    if request.method != 'POST':
        return redirect('core:admin_panel:product_list')
    product = get_object_or_404(Product, pk=pk)
    product.is_active = not product.is_active
    product.save()
    action = 'activado' if product.is_active else 'inactivado'
    messages.success(request, f'Producto {action} correctamente.')
    return redirect('core:admin_panel:product_list')


class ProductDeleteView(StaffRequiredMixin, DeleteView):
    model = Product
    template_name = 'dashboard/product_confirm_delete.html'
    context_object_name = 'product'
    success_url = reverse_lazy('core:admin_panel:product_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Producto eliminado.')
        return super().delete(request, *args, **kwargs)


# --- Reseñas (Reviews) ---

class ReviewListView(StaffRequiredMixin, ListView):
    model = ProductReview
    template_name = 'dashboard/review_list.html'
    context_object_name = 'reviews'
    paginate_by = 25

    def get_queryset(self):
        qs = ProductReview.objects.select_related('product').order_by('-created_at')
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                Q(author_name__icontains=search) |
                Q(author_email__icontains=search) |
                Q(comment__icontains=search) |
                Q(product__name__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'pending':
            qs = qs.filter(is_approved=False)
        elif status == 'approved':
            qs = qs.filter(is_approved=True)
        sort = self.request.GET.get('sort', '-created_at')
        order_map = {
            '-created_at': ['-created_at'],
            'created_at': ['created_at'],
            'rating': ['rating', '-created_at'],
            '-rating': ['-rating', '-created_at'],
            'author': ['author_name', '-created_at'],
            '-author': ['-author_name', '-created_at'],
        }
        return qs.order_by(*order_map.get(sort, ['-created_at']))


# --- Contactos (Formulario /contacto/) ---

class ContactSubmissionListView(StaffRequiredMixin, ListView):
    model = ContactSubmission
    template_name = 'dashboard/contact_submission_list.html'
    context_object_name = 'submissions'
    paginate_by = 30

    def get_queryset(self):
        qs = ContactSubmission.objects.all()
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(email__icontains=search)
                | Q(phone__icontains=search)
                | Q(message__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'unread':
            qs = qs.filter(is_read=False)
        elif status == 'read':
            qs = qs.filter(is_read=True)
        sort = self.request.GET.get('sort', '-created_at')
        order_map = {
            '-created_at': ['-created_at', '-id'],
            'created_at': ['created_at', 'id'],
            'name': ['name', '-created_at'],
            '-name': ['-name', '-created_at'],
        }
        return qs.order_by(*order_map.get(sort, ['-created_at', '-id']))


class ContactSubmissionDetailView(StaffRequiredMixin, DetailView):
    model = ContactSubmission
    template_name = 'dashboard/contact_submission_detail.html'
    context_object_name = 'submission'

    def get(self, request, *args, **kwargs):
        resp = super().get(request, *args, **kwargs)
        obj = self.get_object()
        if obj and not obj.is_read:
            ContactSubmission.objects.filter(pk=obj.pk).update(is_read=True)
        return resp


@_dashboard_required
def contact_submission_export_excel_view(request):
    """Exporta mensajes de contacto a Excel (.xlsx)."""
    from django.utils import timezone

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ModuleNotFoundError:
        messages.error(
            request,
            'Para exportar en Excel debe instalar openpyxl. '
            'Ejecute: pip install openpyxl y reinicie el servidor.',
        )
        return redirect('core:admin_panel:contact_submission_list')

    qs = ContactSubmission.objects.all().order_by('-created_at', '-id')
    timestamp = timezone.now().strftime('%Y%m%d-%H%M')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contactos'
    headers = [
        'Estado',
        'Nombre',
        'Email',
        'Teléfono',
        'Mensaje',
        'IP',
        'Fecha',
    ]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, s in enumerate(qs, 2):
        ws.cell(row=row, column=1, value='Leído' if s.is_read else 'Nuevo')
        ws.cell(row=row, column=2, value=s.name)
        ws.cell(row=row, column=3, value=s.email)
        ws.cell(row=row, column=4, value=s.phone)
        ws.cell(row=row, column=5, value=s.message)
        ws.cell(row=row, column=6, value=s.ip_address or '')
        ws.cell(row=row, column=7, value=s.created_at.strftime('%Y-%m-%d %H:%M'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'contactos-{timestamp}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_dashboard_required
def contact_submission_toggle_read_view(request, pk):
    if request.method != 'POST':
        return redirect('core:admin_panel:contact_submission_detail', pk=pk)
    obj = get_object_or_404(ContactSubmission, pk=pk)
    obj.is_read = not obj.is_read
    obj.save(update_fields=['is_read'])
    return redirect('core:admin_panel:contact_submission_detail', pk=pk)


@_dashboard_required
def contact_submission_delete_view(request, pk):
    if request.method != 'POST':
        return redirect('core:admin_panel:contact_submission_detail', pk=pk)
    obj = get_object_or_404(ContactSubmission, pk=pk)
    obj.delete()
    messages.success(request, 'Mensaje eliminado.')
    return redirect('core:admin_panel:contact_submission_list')


@_dashboard_required
def review_approve_view(request, pk):
    """Autoriza una reseña (la hace visible en la tienda)."""
    if request.method != 'POST':
        return redirect('core:admin_panel:review_list')
    review = get_object_or_404(ProductReview, pk=pk)
    review.is_approved = True
    review.save()
    messages.success(request, f'Reseña de {review.author_name} autorizada.')
    next_url = _safe_next_url(
        request,
        request.GET.get('next') or request.POST.get('next'),
        reverse('core:admin_panel:review_list'),
    )
    return redirect(next_url)


@_dashboard_required
def review_reject_view(request, pk):
    """Rechaza una reseña (deja de mostrarla)."""
    if request.method != 'POST':
        return redirect('core:admin_panel:review_list')
    review = get_object_or_404(ProductReview, pk=pk)
    review.is_approved = False
    review.save()
    messages.success(request, f'Reseña de {review.author_name} rechazada.')
    next_url = _safe_next_url(
        request,
        request.GET.get('next') or request.POST.get('next'),
        reverse('core:admin_panel:review_list'),
    )
    return redirect(next_url)


@_dashboard_required
def review_delete_view(request, pk):
    """Elimina una reseña (spam, etc.)."""
    if request.method != 'POST':
        return redirect('core:admin_panel:review_list')
    review = get_object_or_404(ProductReview, pk=pk)
    author = review.author_name
    review.delete()
    messages.success(request, f'Reseña de {author} eliminada.')
    return redirect('core:admin_panel:review_list')


# --- Clientes ---

class CustomerListView(StaffRequiredMixin, ListView):
    model = User
    template_name = 'dashboard/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 25

    def get_queryset(self):
        qs = User.objects.annotate(
            orders_count=models.Count('orders')
        )
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                models.Q(username__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search) |
                models.Q(phone__icontains=search) |
                models.Q(document_number__icontains=search)
            )
        role = self.request.GET.get('role')
        if role:
            qs = qs.filter(role=role)
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        customer_type = self.request.GET.get('customer_type')
        if customer_type:
            qs = qs.filter(customer_type=customer_type)
        has_orders = self.request.GET.get('has_orders')
        if has_orders == 'yes':
            qs = qs.filter(orders_count__gt=0)
        elif has_orders == 'no':
            qs = qs.filter(orders_count=0)
        sort = self.request.GET.get('sort', '-date_joined')
        order_map = {
            '-date_joined': ['-date_joined'],
            'date_joined': ['date_joined'],
            'name': ['first_name', 'last_name', 'email'],
            '-name': ['-first_name', '-last_name', 'email'],
            'email': ['email'],
            '-email': ['-email'],
            'orders': ['orders_count', '-date_joined'],
            '-orders': ['-orders_count', '-date_joined'],
        }
        qs = qs.order_by(*order_map.get(sort, ['-date_joined']))
        return qs


class CustomerCreateView(StaffRequiredMixin, CreateView):
    model = User
    form_class = CustomerCreateForm
    template_name = 'dashboard/customer_form.html'
    success_url = reverse_lazy('core:admin_panel:customer_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import Country
        import json
        ctx['geo_countries_json'] = json.dumps([{'id': c.id, 'name': c.name} for c in Country.objects.all().order_by('name')])
        ctx['geo_states_url'] = reverse('core:geo_states')
        ctx['geo_cities_url'] = reverse('core:geo_cities')
        ctx['initial_state'] = ''
        ctx['initial_city'] = ''
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Cliente creado correctamente.')
        return super().form_valid(form)


@_dashboard_required
def customer_detail_view(request, pk):
    """Detalle de cliente con sus pedidos."""
    from django.db.models import Sum
    customer = get_object_or_404(User.objects.prefetch_related('orders'), pk=pk)
    orders = customer.orders.all().order_by('-created_at')[:20]
    total_spent = customer.orders.filter(
        status__in=['completed', 'processing']
    ).aggregate(Sum('total'))['total__sum'] or 0
    return render(request, 'dashboard/customer_detail.html', {
        'customer': customer,
        'orders': orders,
        'total_spent': total_spent,
    })


class CustomerUpdateView(StaffRequiredMixin, UpdateView):
    model = User
    form_class = CustomerForm
    template_name = 'dashboard/customer_form.html'
    context_object_name = 'customer'
    success_url = reverse_lazy('core:admin_panel:customer_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import Country
        import json
        ctx['geo_countries_json'] = json.dumps([{'id': c.id, 'name': c.name} for c in Country.objects.all().order_by('name')])
        ctx['geo_states_url'] = reverse('core:geo_states')
        ctx['geo_cities_url'] = reverse('core:geo_cities')
        customer = ctx.get('customer')
        ctx['initial_state'] = getattr(customer, 'state', '') or ''
        ctx['initial_city'] = getattr(customer, 'city', '') or ''
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Cliente actualizado correctamente.')
        return super().form_valid(form)


# --- Newsletter ---
class NewsletterSubscriberListView(StaffRequiredMixin, ListView):
    model = NewsletterSubscriber
    template_name = 'dashboard/newsletter_list.html'
    context_object_name = 'subscribers'
    paginate_by = 30

    def get_queryset(self):
        qs = NewsletterSubscriber.objects.all()
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(email__icontains=search)
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        sort = self.request.GET.get('sort', '-created_at')
        order_map = {
            '-created_at': ['-created_at'],
            'created_at': ['created_at'],
            'email': ['email'],
            '-email': ['-email'],
        }
        return qs.order_by(*order_map.get(sort, ['-created_at']))


@_dashboard_required
def newsletter_export_excel_view(request):
    """Exporta suscriptores de newsletter a Excel (.xlsx)."""
    from django.utils import timezone

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ModuleNotFoundError:
        messages.error(
            request,
            'Para exportar en Excel debe instalar openpyxl. '
            'Ejecute: pip install openpyxl y reinicie el servidor.',
        )
        return redirect('core:admin_panel:newsletter_list')

    qs = NewsletterSubscriber.objects.all().order_by('-created_at')
    timestamp = timezone.now().strftime('%Y%m%d-%H%M')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Newsletter'
    headers = ['Email', 'Estado', 'Origen', 'Suscrito en', 'Actualizado en']
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, sub in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=sub.email)
        ws.cell(row=row, column=2, value='Activo' if sub.is_active else 'Inactivo')
        ws.cell(row=row, column=3, value=sub.source or 'footer')
        ws.cell(row=row, column=4, value=sub.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=5, value=sub.updated_at.strftime('%Y-%m-%d %H:%M'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'newsletter-suscriptores-{timestamp}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_dashboard_required
def newsletter_toggle_active_view(request, pk):
    if request.method != 'POST':
        return redirect('core:admin_panel:newsletter_list')
    sub = get_object_or_404(NewsletterSubscriber, pk=pk)
    sub.is_active = not sub.is_active
    sub.save(update_fields=['is_active', 'updated_at'])
    action = 'activado' if sub.is_active else 'inactivado'
    messages.success(request, f'Suscriptor {sub.email} {action} correctamente.')
    return redirect('core:admin_panel:newsletter_list')


@_dashboard_required
def newsletter_delete_view(request, pk):
    if request.method != 'POST':
        return redirect('core:admin_panel:newsletter_list')
    sub = get_object_or_404(NewsletterSubscriber, pk=pk)
    email = sub.email
    sub.delete()
    messages.success(request, f'Suscriptor {email} eliminado.')
    return redirect('core:admin_panel:newsletter_list')


# --- Pedidos ---

@method_decorator(never_cache, name='dispatch')
class OrderListView(StaffRequiredMixin, ListView):
    model = Order
    template_name = 'dashboard/order_list.html'
    context_object_name = 'orders'
    paginate_by = 25

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related('user')
            .prefetch_related('wompi_transactions')
        )
        tab = (self.request.GET.get('tab') or 'activos').strip().lower()
        if tab == 'cancelados':
            qs = qs.filter(status='cancelled')
        elif tab == 'completados':
            qs = qs.filter(status='completed')
        else:
            qs = qs.exclude(status__in=['cancelled', 'completed'])
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                models.Q(order_number__icontains=search) |
                models.Q(billing_email__icontains=search) |
                models.Q(billing_first_name__icontains=search) |
                models.Q(billing_last_name__icontains=search) |
                models.Q(billing_phone__icontains=search)
            )
        if tab in ('activos', 'completados'):
            status = self.request.GET.get('status')
            if status:
                qs = qs.filter(status=status)
            payment_status = self.request.GET.get('payment_status')
            if payment_status:
                qs = qs.filter(payment_status=payment_status)
        date_from = (self.request.GET.get('date_from') or '').strip()
        if date_from:
            try:
                from datetime import datetime
                dt = datetime.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__gte=dt)
            except ValueError:
                pass
        date_to = (self.request.GET.get('date_to') or '').strip()
        if date_to:
            try:
                from datetime import datetime
                dt = datetime.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(created_at__date__lte=dt)
            except ValueError:
                pass
        filter_state_id = self.request.GET.get('filter_state')
        filter_city_id = self.request.GET.get('filter_city')
        if filter_city_id:
            try:
                city = City.objects.select_related('state').get(pk=filter_city_id)
                qs = qs.filter(
                    billing_state__iexact=city.state.name,
                    billing_city__iexact=city.name,
                )
            except (City.DoesNotExist, ValueError):
                pass
        elif filter_state_id:
            try:
                state = State.objects.get(pk=filter_state_id)
                qs = qs.filter(billing_state__iexact=state.name)
            except (State.DoesNotExist, ValueError):
                pass
        sort = self.request.GET.get('sort', '-created_at')
        order_map = {
            '-created_at': ['-created_at'],
            'created_at': ['created_at'],
            'total': ['total', '-created_at'],
            '-total': ['-total', '-created_at'],
        }
        return qs.order_by(*order_map.get(sort, ['-created_at']))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tab = (self.request.GET.get('tab') or 'activos').strip().lower()
        context['orders_tab'] = tab if tab in ('cancelados', 'completados') else 'activos'
        context['orders_cancelled_count'] = Order.objects.filter(status='cancelled').count()
        context['orders_completed_count'] = Order.objects.filter(status='completed').count()
        context['geo_states_url'] = reverse('core:geo_states')
        context['geo_cities_url'] = reverse('core:geo_cities')
        context['filter_countries'] = Country.objects.all().order_by('name')
        filter_country_id = self.request.GET.get('filter_country')
        if filter_country_id:
            context['filter_states'] = State.objects.filter(country_id=filter_country_id).order_by('name')
            context['filter_country_id'] = filter_country_id
        else:
            context['filter_states'] = State.objects.none()
            context['filter_country_id'] = ''
        filter_state_id = self.request.GET.get('filter_state')
        if filter_state_id:
            context['filter_cities'] = City.objects.filter(state_id=filter_state_id).order_by('name')
            context['filter_state_id'] = filter_state_id
        else:
            context['filter_cities'] = City.objects.none()
            context['filter_state_id'] = ''
        context['filter_city_id'] = self.request.GET.get('filter_city') or ''
        orders = context.get(self.context_object_name, [])
        for order in orders:
            if order.payment_status != 'pending':
                continue
            latest_tx = order.wompi_transactions.first()
            if latest_tx and latest_tx.status == 'APPROVED':
                order.payment_status = 'paid'
                order.save(update_fields=['payment_status', 'updated_at'])
        return context


@_dashboard_required
def order_detail_view(request, pk):
    """Detalle y actualización de pedido."""
    order = get_object_or_404(
        Order.objects.prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product').prefetch_related('product__images')),
            'wompi_transactions',
            Prefetch('order_notes', queryset=OrderNote.objects.select_related('created_by').order_by('-created_at')),
        ),
        pk=pk,
    )
    if request.method == 'POST':
        note_content = (request.POST.get('note_content') or '').strip()
        if request.POST.get('add_internal_note'):
            if note_content:
                OrderNote.objects.create(
                    order=order,
                    note_type=OrderNote.NOTE_TYPE_INTERNAL,
                    content=note_content,
                    created_by=request.user if request.user.is_authenticated else None,
                )
                messages.success(request, 'Nota interna guardada.')
            else:
                messages.warning(request, 'Escribe el contenido de la nota.')
            return redirect('core:admin_panel:order_detail', pk=order.pk)
        if request.POST.get('add_client_note'):
            if note_content:
                from apps.core.emails import notify_order_note_to_customer
                OrderNote.objects.create(
                    order=order,
                    note_type=OrderNote.NOTE_TYPE_CLIENT,
                    content=note_content,
                    created_by=request.user if request.user.is_authenticated else None,
                )
                try:
                    notify_order_note_to_customer(order, note_content)
                except Exception:
                    import logging
                    logging.getLogger(__name__).exception(
                        "Error enviando nota al cliente para pedido %s", order.order_number
                    )
                messages.success(request, 'Nota guardada y enviada por email al cliente.')
            else:
                messages.warning(request, 'Escribe el contenido de la nota para el cliente.')
            return redirect('core:admin_panel:order_detail', pk=order.pk)
        form = OrderStatusForm(request.POST, instance=order)
        if form.is_valid():
            import logging
            from apps.core.emails import notify_order_status_changed

            log = logging.getLogger(__name__)
            old_status = order.status
            old_payment_status = order.payment_status
            new_status = form.cleaned_data["status"]
            new_payment_status = form.cleaned_data["payment_status"]
            form.save()
            status_changed = old_status != new_status or old_payment_status != new_payment_status

            if status_changed:
                if order.billing_email:
                    log.info(
                        "Enviando notificación de cambio de estado pedido %s a %s",
                        order.order_number,
                        order.billing_email,
                    )
                    try:
                        notify_order_status_changed(order)
                        messages.success(request, 'Pedido actualizado. Se ha enviado notificación al cliente.')
                    except Exception:
                        log.exception(
                            "Error enviando notificación de cambio de estado para pedido %s",
                            order.order_number,
                        )
                        messages.warning(
                            request,
                            'Pedido actualizado, pero no se pudo enviar el correo al cliente. Revisa los logs.',
                        )
                else:
                    messages.warning(
                        request,
                        'Pedido actualizado. No se envió correo: el pedido no tiene email de facturación.',
                    )
            else:
                messages.success(request, 'Pedido actualizado. No hubo cambio de estado o pago.')
            return redirect('core:admin_panel:order_detail', pk=order.pk)
    else:
        form = OrderStatusForm(instance=order)

    wompi_tx = order.wompi_transactions.order_by('-created_at').first()
    return render(request, 'dashboard/order_detail.html', {
        'order': order,
        'form': form,
        'wompi_tx': wompi_tx,
    })


# --- Cupones ---

class CouponListView(StaffRequiredMixin, ListView):
    model = Coupon
    template_name = 'dashboard/coupon_list.html'
    context_object_name = 'coupons'
    paginate_by = 20

    def get_queryset(self):
        qs = Coupon.objects.all()
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(code__icontains=search)
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        discount_type = self.request.GET.get('discount_type')
        if discount_type:
            qs = qs.filter(discount_type=discount_type)
        sort = self.request.GET.get('sort', '-created_at')
        order_map = {
            '-created_at': ['-created_at'],
            'created_at': ['created_at'],
            'code': ['code'],
            '-code': ['-code'],
        }
        return qs.order_by(*order_map.get(sort, ['-created_at']))


class CouponCreateView(StaffRequiredMixin, CreateView):
    model = Coupon
    form_class = CouponForm
    template_name = 'dashboard/coupon_form.html'
    success_url = reverse_lazy('core:admin_panel:coupon_list')

    def form_valid(self, form):
        code = (form.cleaned_data.get('code') or '').strip()
        if not code:
            form.instance.code = self._generate_unique_coupon_code()
        messages.success(self.request, 'Cupón creado correctamente.')
        return super().form_valid(form)

    def _generate_unique_coupon_code(self):
        import random
        import string
        for _ in range(100):
            code = 'CUP' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not Coupon.objects.filter(code=code).exists():
                return code
        from django.utils import timezone
        return 'CUP' + timezone.now().strftime('%Y%m%d%H%M') + ''.join(random.choices(string.digits, k=4))


class CouponUpdateView(StaffRequiredMixin, UpdateView):
    model = Coupon
    form_class = CouponForm
    template_name = 'dashboard/coupon_form.html'
    context_object_name = 'coupon'
    success_url = reverse_lazy('core:admin_panel:coupon_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cupón actualizado correctamente.')
        return super().form_valid(form)


class CouponDeleteView(StaffRequiredMixin, DeleteView):
    model = Coupon
    template_name = 'dashboard/coupon_confirm_delete.html'
    context_object_name = 'coupon'
    success_url = reverse_lazy('core:admin_panel:coupon_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Cupón eliminado.')
        return super().delete(request, *args, **kwargs)


# --- Precios de envío por ciudad ---

class ShippingPriceListView(StaffRequiredMixin, ListView):
    model = ShippingPrice
    template_name = 'dashboard/shipping_price_list.html'
    context_object_name = 'shipping_prices'
    paginate_by = 25

    def get_queryset(self):
        qs = ShippingPrice.objects.select_related(
            'city', 'city__state', 'city__state__country'
        )
        search = (self.request.GET.get('q') or '').strip()
        if search:
            qs = qs.filter(
                Q(city__name__icontains=search) |
                Q(city__state__name__icontains=search) |
                Q(city__state__country__name__icontains=search)
            )
        status = self.request.GET.get('status')
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        country_id = self.request.GET.get('country')
        if country_id:
            qs = qs.filter(city__state__country_id=country_id)
        state_id = self.request.GET.get('state')
        if state_id:
            qs = qs.filter(city__state_id=state_id)
        sort = self.request.GET.get('sort', 'location')
        order_map = {
            'location': ['city__state__name', 'city__name'],
            '-location': ['-city__state__name', '-city__name'],
            'city': ['city__name', 'city__state__name'],
            '-city': ['-city__name', 'city__state__name'],
            'price': ['price', 'city__name'],
            '-price': ['-price', 'city__name'],
            'days': ['delivery_days_min', 'city__name'],
            '-days': ['-delivery_days_max', 'city__name'],
        }
        qs = qs.order_by(*order_map.get(sort, order_map['location']))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import Country, State

        ctx['free_shipping_form'] = ShippingFreeRuleForm(instance=SiteSettings.get())
        country_ids = ShippingPrice.objects.values_list(
            'city__state__country_id', flat=True
        ).distinct()
        ctx['filter_countries'] = Country.objects.filter(
            id__in=country_ids
        ).order_by('name')
        state_ids = ShippingPrice.objects.values_list(
            'city__state_id', flat=True
        ).distinct()
        ctx['filter_states'] = State.objects.filter(
            id__in=state_ids
        ).select_related('country').order_by('country__name', 'name')
        return ctx


@_dashboard_required
def shipping_free_rule_update_view(request):
    """Actualiza el monto mínimo para envío gratis."""
    if request.method != 'POST':
        return redirect('core:admin_panel:shipping_price_list')
    settings_obj = SiteSettings.get()
    form = ShippingFreeRuleForm(request.POST, instance=settings_obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Configuración de envío gratis guardada.')
    else:
        messages.error(request, 'No se pudo guardar el envío gratis. Verifica el valor ingresado.')
    return redirect('core:admin_panel:shipping_price_list')


@_dashboard_required
def shipping_price_export_excel_view(request):
    """Exporta todos los precios de envío a Excel (.xlsx). Requiere openpyxl."""
    from django.utils import timezone

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ModuleNotFoundError:
        messages.error(
            request,
            'Para exportar en Excel debe instalar openpyxl. '
            'En la misma terminal donde ejecuta el servidor (runserver), ejecute: pip install openpyxl '
            'y luego reinicie el servidor.'
        )
        return redirect('core:admin_panel:shipping_price_list')

    qs = ShippingPrice.objects.select_related('city', 'city__state', 'city__state__country').order_by('city__state__name', 'city__name')
    timestamp = timezone.now().strftime('%Y%m%d-%H%M')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Precios de envío'
    headers = ['Ciudad', 'Departamento', 'País', 'Precio', 'Días mín', 'Días máx', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row, sp in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=sp.city.name)
        ws.cell(row=row, column=2, value=sp.city.state.name)
        ws.cell(row=row, column=3, value=sp.city.state.country.name)
        ws.cell(row=row, column=4, value=float(sp.price))
        ws.cell(row=row, column=5, value=sp.delivery_days_min)
        ws.cell(row=row, column=6, value=sp.delivery_days_max)
        ws.cell(row=row, column=7, value='Activo' if sp.is_active else 'Inactivo')
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'precios-envio-{timestamp}.xlsx'
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_dashboard_required
def shipping_price_load_all_colombia_view(request):
    """Crea precio de envío (0, 0-0 días) para todas las ciudades de Colombia que aún no tengan uno."""
    if request.method != 'POST':
        return redirect('core:admin_panel:shipping_price_list')
    try:
        colombia = Country.objects.get(name='Colombia')
    except Country.DoesNotExist:
        messages.error(request, 'No existe el país Colombia. Ejecute: python manage.py load_colombia_geo')
        return redirect('core:admin_panel:shipping_price_list')
    cities = City.objects.filter(state__country=colombia).select_related('state')
    created = 0
    for city in cities:
        _, created_flag = ShippingPrice.objects.get_or_create(
            city=city,
            defaults={
                'price': 0,
                'delivery_days_min': 0,
                'delivery_days_max': 0,
                'is_active': True,
            }
        )
        if created_flag:
            created += 1
    messages.success(request, f'Se agregaron {created} ciudades de Colombia con precio 0 y días 0-0. ({cities.count() - created} ya existían.)')
    return redirect('core:admin_panel:shipping_price_list')


@_dashboard_required
def shipping_price_import_excel_view(request):
    """Importa precios de envío desde Excel (.xlsx)."""
    if request.method == 'GET':
        return render(request, 'dashboard/shipping_price_import.html')

    file = request.FILES.get('excel_file')
    if not file:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('core:admin_panel:shipping_price_import_excel')

    if not file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser Excel (.xlsx o .xls).')
        return redirect('core:admin_panel:shipping_price_import_excel')

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        messages.error(
            request,
            'Para importar Excel instale openpyxl: pip install openpyxl'
        )
        return redirect('core:admin_panel:shipping_price_list')

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
    except Exception as e:
        messages.error(request, f'No se pudo leer el archivo Excel: {e}')
        return redirect('core:admin_panel:shipping_price_import_excel')

    ws = wb['Precios de envío'] if 'Precios de envío' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'El archivo está vacío o no tiene filas.')
        wb.close()
        return redirect('core:admin_panel:shipping_price_import_excel')

    header_row = [str(c).strip() if c is not None else '' for c in rows[0]]
    col_map = {}
    aliases = {
        'ciudad': 'ciudad', 'departamento': 'departamento', 'pais': 'pais',
        'país': 'pais', 'precio': 'precio',
        'días mín': 'dias_min', 'dias mín': 'dias_min', 'días min': 'dias_min', 'dias min': 'dias_min',
        'dias mn': 'dias_min', 'días mn': 'dias_min',
        'días máx': 'dias_max', 'dias máx': 'dias_max', 'días max': 'dias_max', 'dias max': 'dias_max',
        'dias mx': 'dias_max', 'días mx': 'dias_max', 'estado': 'estado',
    }
    for i, h in enumerate(header_row):
        h_clean = h.lower().strip()
        key = aliases.get(h_clean)
        if key and key not in col_map:
            col_map[key] = i
        if not key and ('min' in h_clean or 'mín' in h_clean or 'mn' in h_clean) and ('dia' in h_clean or 'días' in h_clean or 'dias' in h_clean):
            col_map.setdefault('dias_min', i)
        if not key and ('max' in h_clean or 'máx' in h_clean or 'mx' in h_clean) and ('dia' in h_clean or 'días' in h_clean or 'dias' in h_clean):
            col_map.setdefault('dias_max', i)

    if len(header_row) >= 6 and 'dias_min' not in col_map:
        col_map['dias_min'] = 4
    if len(header_row) >= 7 and 'dias_max' not in col_map:
        col_map['dias_max'] = 5

    if 'ciudad' not in col_map or 'precio' not in col_map:
        messages.error(
            request,
            'El Excel debe tener columnas: Ciudad, Departamento, País, Precio, Días mín, Días máx, Estado'
        )
        wb.close()
        return redirect('core:admin_panel:shipping_price_import_excel')

    created = 0
    updated = 0
    not_found = []

    def _val(row, key, default=None):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        if v is None or (isinstance(v, str) and not v.strip()):
            return default
        return v

    for row in rows[1:]:
        row = list(row) if row else []
        ciudad = str(_val(row, 'ciudad', '') or '').strip()
        if not ciudad:
            continue
        departamento = str(_val(row, 'departamento', '') or '').strip()
        pais = str(_val(row, 'pais', '') or '').strip() or 'Colombia'
        try:
            price_val = _val(row, 'precio', 0)
            price = float(price_val) if price_val is not None else 0
        except (TypeError, ValueError):
            price = 0
        try:
            dmin = int(_val(row, 'dias_min', 0) or 0)
        except (TypeError, ValueError):
            dmin = 0
        try:
            dmax = int(_val(row, 'dias_max', 0) or 0)
        except (TypeError, ValueError):
            dmax = 0
        estado = str(_val(row, 'estado', 'Activo') or 'Activo').strip().lower()
        is_active = estado.startswith('activo')

        city = City.objects.filter(
            name__iexact=ciudad,
            state__name__iexact=departamento,
            state__country__name__iexact=pais,
        ).select_related('state', 'state__country').first()

        if not city:
            not_found.append(f'{ciudad}, {departamento}, {pais}')
            continue

        _, was_created = ShippingPrice.objects.update_or_create(
            city=city,
            defaults={
                'price': price,
                'delivery_days_min': max(0, dmin),
                'delivery_days_max': max(0, dmax),
                'is_active': is_active,
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    wb.close()

    msg_parts = []
    if created:
        msg_parts.append(f'{created} creados')
    if updated:
        msg_parts.append(f'{updated} actualizados')
    if not msg_parts:
        msg_parts.append('0 registros procesados')
    messages.success(request, 'Importación completada: ' + ', '.join(msg_parts))
    if not_found:
        samples = not_found[:5]
        extra = f' ({len(not_found)} total)' if len(not_found) > 5 else ''
        messages.warning(
            request,
            f'Ciudades no encontradas en la base de datos{extra}: ' + '; '.join(samples)
        )
    return redirect('core:admin_panel:shipping_price_list')


class ShippingPriceCreateView(StaffRequiredMixin, CreateView):
    model = ShippingPrice
    form_class = ShippingPriceForm
    template_name = 'dashboard/shipping_price_form.html'
    success_url = reverse_lazy('core:admin_panel:shipping_price_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['geo_cities_url'] = reverse('core:geo_cities')
        ctx['initial_state_id'] = ''
        ctx['initial_city_id'] = ''
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Precio de envío agregado correctamente.')
        return super().form_valid(form)


class ShippingPriceUpdateView(StaffRequiredMixin, UpdateView):
    model = ShippingPrice
    form_class = ShippingPriceForm
    template_name = 'dashboard/shipping_price_form.html'
    context_object_name = 'shipping_price'
    success_url = reverse_lazy('core:admin_panel:shipping_price_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['geo_cities_url'] = reverse('core:geo_cities')
        obj = ctx.get('shipping_price')
        if obj and obj.city_id:
            ctx['initial_state_id'] = obj.city.state_id
            ctx['initial_city_id'] = obj.city_id
        else:
            ctx['initial_state_id'] = ''
            ctx['initial_city_id'] = ''
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Precio de envío actualizado.')
        return super().form_valid(form)


class ShippingPriceDeleteView(StaffRequiredMixin, DeleteView):
    model = ShippingPrice
    template_name = 'dashboard/shipping_price_confirm_delete.html'
    context_object_name = 'shipping_price'
    success_url = reverse_lazy('core:admin_panel:shipping_price_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Precio de envío eliminado.')
        return super().delete(request, *args, **kwargs)


# --- Configuración del sitio ---

class SiteSettingsUpdateView(StaffRequiredMixin, UpdateView):
    model = SiteSettings
    form_class = SiteSettingsForm
    template_name = 'dashboard/config_form.html'
    context_object_name = 'settings'
    success_url = reverse_lazy('core:admin_panel:config')

    def get_object(self, queryset=None):
        return SiteSettings.get()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import Country
        import json
        ctx['geo_countries_json'] = json.dumps([{'id': c.id, 'name': c.name} for c in Country.objects.all().order_by('name')])
        ctx['geo_states_url'] = reverse('core:geo_states')
        ctx['geo_cities_url'] = reverse('core:geo_cities')
        obj = ctx.get('settings')
        ctx['initial_state'] = getattr(obj, 'state', '') or ''
        ctx['initial_city'] = getattr(obj, 'city', '') or ''
        return ctx

    def form_valid(self, form):
        messages.success(self.request, 'Configuración guardada correctamente.')
        return super().form_valid(form)


# --- Secciones del Home ---

class HomeSectionsConfigView(StaffRequiredMixin, ListView):
    """Vista principal de administración de secciones del home."""
    model = HomeSection
    template_name = 'dashboard/home_sections_config.html'
    context_object_name = 'sections'

    def get_queryset(self):
        return HomeSection.objects.all().order_by('order')

    def get_context_data(self, **kwargs):
        from django.forms import modelformset_factory
        from .forms import HomeSectionForm
        ctx = super().get_context_data(**kwargs)
        FormSet = modelformset_factory(HomeSection, form=HomeSectionForm, extra=0)
        ctx['formset'] = kwargs.get('formset') or FormSet(queryset=self.get_queryset())
        ctx['hero_slides_count'] = HomeHeroSlide.objects.count()
        ctx['brands_count'] = Brand.objects.count()
        ctx['testimonials_count'] = HomeTestimonial.objects.count()
        return ctx

    def post(self, request, *args, **kwargs):
        from django.forms import modelformset_factory
        from .forms import HomeSectionForm
        self.object_list = self.get_queryset()
        FormSet = modelformset_factory(HomeSection, form=HomeSectionForm, extra=0)
        formset = FormSet(request.POST, queryset=self.get_queryset())
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Configuración de secciones guardada.')
            return redirect('core:admin_panel:home_sections')
        return self.render_to_response(self.get_context_data(formset=formset))


class HomeHeroSlideListView(StaffRequiredMixin, ListView):
    model = HomeHeroSlide
    template_name = 'dashboard/home_hero_list.html'
    context_object_name = 'slides'


class HomeHeroSlideCreateView(StaffRequiredMixin, CreateView):
    model = HomeHeroSlide
    form_class = HomeHeroSlideForm
    template_name = 'dashboard/home_hero_form.html'
    success_url = reverse_lazy('core:admin_panel:home_hero_list')

    def form_valid(self, form):
        messages.success(self.request, 'Slide creado.')
        return super().form_valid(form)


class HomeHeroSlideUpdateView(StaffRequiredMixin, UpdateView):
    model = HomeHeroSlide
    form_class = HomeHeroSlideForm
    template_name = 'dashboard/home_hero_form.html'
    context_object_name = 'slide'
    success_url = reverse_lazy('core:admin_panel:home_hero_list')

    def form_valid(self, form):
        messages.success(self.request, 'Slide actualizado.')
        return super().form_valid(form)


class HomeHeroSlideDeleteView(StaffRequiredMixin, DeleteView):
    model = HomeHeroSlide
    template_name = 'dashboard/home_hero_confirm_delete.html'
    context_object_name = 'slide'
    success_url = reverse_lazy('core:admin_panel:home_hero_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Slide eliminado.')
        return super().delete(request, *args, **kwargs)


class HomeAboutBlockUpdateView(StaffRequiredMixin, UpdateView):
    model = HomeAboutBlock
    form_class = HomeAboutBlockForm
    template_name = 'dashboard/home_about_form.html'
    context_object_name = 'block'
    success_url = reverse_lazy('core:admin_panel:home_sections')

    def get_object(self, queryset=None):
        return HomeAboutBlock.get()

    def form_valid(self, form):
        messages.success(self.request, 'Bloque sobre nosotros actualizado.')
        return super().form_valid(form)


class HomeMeatCategoryBlockUpdateView(StaffRequiredMixin, UpdateView):
    model = HomeMeatCategoryBlock
    form_class = HomeMeatCategoryBlockForm
    template_name = 'dashboard/home_meat_category_form.html'
    context_object_name = 'block'
    success_url = reverse_lazy('core:admin_panel:home_sections')

    def get_object(self, queryset=None):
        return HomeMeatCategoryBlock.get()

    def form_valid(self, form):
        messages.success(self.request, 'Sección categorías actualizada.')
        return super().form_valid(form)


class HomeBrandBlockUpdateView(StaffRequiredMixin, UpdateView):
    """Configuración de la sección Marcas (imagen de fondo)."""
    model = HomeBrandBlock
    form_class = HomeBrandBlockForm
    template_name = 'dashboard/home_brand_block_form.html'
    context_object_name = 'block'
    success_url = reverse_lazy('core:admin_panel:home_brand_list')

    def get_object(self, queryset=None):
        return HomeBrandBlock.get()

    def form_valid(self, form):
        messages.success(self.request, 'Imagen de fondo guardada.')
        return super().form_valid(form)


class HomeBrandListView(StaffRequiredMixin, ListView):
    """Lista las marcas del catálogo de productos (Brand)."""
    model = Brand
    template_name = 'dashboard/home_brand_list.html'
    context_object_name = 'brands'
    queryset = Brand.objects.all().order_by('order', 'name')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['brand_block'] = HomeBrandBlock.get()
        return ctx


class HomeBrandCreateView(StaffRequiredMixin, CreateView):
    model = HomeBrand
    form_class = HomeBrandForm
    template_name = 'dashboard/home_brand_form.html'
    success_url = reverse_lazy('core:admin_panel:home_brand_list')

    def form_valid(self, form):
        messages.success(self.request, 'Marca creada.')
        return super().form_valid(form)


class HomeBrandUpdateView(StaffRequiredMixin, UpdateView):
    model = HomeBrand
    form_class = HomeBrandForm
    template_name = 'dashboard/home_brand_form.html'
    context_object_name = 'brand'
    success_url = reverse_lazy('core:admin_panel:home_brand_list')

    def form_valid(self, form):
        messages.success(self.request, 'Marca actualizada.')
        return super().form_valid(form)


class HomeBrandDeleteView(StaffRequiredMixin, DeleteView):
    model = HomeBrand
    template_name = 'dashboard/home_brand_confirm_delete.html'
    context_object_name = 'brand'
    success_url = reverse_lazy('core:admin_panel:home_brand_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Marca eliminada.')
        return super().delete(request, *args, **kwargs)


class HomeTestimonialListView(StaffRequiredMixin, ListView):
    model = HomeTestimonial
    template_name = 'dashboard/home_testimonial_list.html'
    context_object_name = 'testimonials'


class HomeTestimonialCreateView(StaffRequiredMixin, CreateView):
    model = HomeTestimonial
    form_class = HomeTestimonialForm
    template_name = 'dashboard/home_testimonial_form.html'
    success_url = reverse_lazy('core:admin_panel:home_testimonial_list')

    def form_valid(self, form):
        messages.success(self.request, 'Testimonio creado.')
        return super().form_valid(form)


class HomeTestimonialUpdateView(StaffRequiredMixin, UpdateView):
    model = HomeTestimonial
    form_class = HomeTestimonialForm
    template_name = 'dashboard/home_testimonial_form.html'
    context_object_name = 'testimonial'
    success_url = reverse_lazy('core:admin_panel:home_testimonial_list')

    def form_valid(self, form):
        messages.success(self.request, 'Testimonio actualizado.')
        return super().form_valid(form)


class HomeTestimonialDeleteView(StaffRequiredMixin, DeleteView):
    model = HomeTestimonial
    template_name = 'dashboard/home_testimonial_confirm_delete.html'
    context_object_name = 'testimonial'
    success_url = reverse_lazy('core:admin_panel:home_testimonial_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Testimonio eliminado.')
        return super().delete(request, *args, **kwargs)


class HomePopupAnnouncementUpdateView(StaffRequiredMixin, UpdateView):
    model = HomePopupAnnouncement
    form_class = HomePopupAnnouncementForm
    template_name = 'dashboard/home_popup_form.html'
    context_object_name = 'popup'
    success_url = reverse_lazy('core:admin_panel:home_sections')

    def get_object(self, queryset=None):
        return HomePopupAnnouncement.get()

    def form_valid(self, form):
        messages.success(self.request, 'Popup del home actualizado.')
        return super().form_valid(form)
